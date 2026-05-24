import openpyxl, json
wb = openpyxl.load_workbook("Grifols_Model_v8.xlsx", data_only=True, read_only=True)
CF = wb["CF"]
def num(ws,r,c):
    v=ws.cell(r,c).value
    return float(v) if isinstance(v,(int,float)) else None
# year -> IS quarter cols / BS year-end col
ISCOLS={2024:[3,4,5,6],2025:[7,8,9,10],2026:[11,12,13,14],2027:[15,16,17,18],2028:[19,20,21,22],2029:[23,24,25,26],2030:[27,28,29,30]}
YE={2024:6,2025:10,2026:14,2027:18,2028:22,2029:26,2030:30}
def fy(ws,r,cols):
    vals=[num(ws,r,c) for c in cols]; vals=[v for v in vals if v is not None]
    return round(sum(vals),1) if vals else None

print("=== CF SHEET LABELS (col A) ===")
for r in range(1, CF.max_row+1):
    v=CF.cell(r,1).value
    if v is not None:
        s=str(v); s=s if len(s)<=55 else s[:55]
        print(f"  {r}: {s}")

print("\n=== CF — FY2024 / FY2025 (sum quarters), key lines ===")
# we'll print all numeric rows for 2024 & 2025
for r in range(8, CF.max_row+1):
    lab=CF.cell(r,1).value
    if lab is None: continue
    v24=fy(CF,r,ISCOLS[2024]); v25=fy(CF,r,ISCOLS[2025])
    if v24 is None and v25 is None: continue
    s=str(lab); s=s if len(s)<=42 else s[:42]
    print(f"  r{r:<3} {s:<44} FY24 {('' if v24 is None else format(v24,',.1f')):>11}  FY25 {('' if v25 is None else format(v25,',.1f')):>11}")
wb.close()
print("\nDONE")
