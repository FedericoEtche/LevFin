import openpyxl, sys

def dump_headers(ws, max_rows=7, max_cols=40):
    print(f"  dims: max_row={ws.max_row} max_col={ws.max_column}")
    for r in range(1, min(max_rows, ws.max_row)+1):
        cells=[]
        for c in range(1, min(max_cols, ws.max_column)+1):
            v=ws.cell(r,c).value
            if v is not None:
                s=str(v)
                if len(s)>22: s=s[:22]
                cells.append(f"{openpyxl.utils.get_column_letter(c)}{r}={s}")
        if cells:
            print("   ", " | ".join(cells))

def dump_colA(ws, col=1, max_rows=120):
    print(f"  --- labels col {openpyxl.utils.get_column_letter(col)} ---")
    for r in range(1, min(max_rows, ws.max_row)+1):
        v=ws.cell(r,col).value
        if v is not None:
            s=str(v)
            if len(s)>50: s=s[:50]
            print(f"    {r}: {s}")

print("="*70)
print("MODEL: Grifols_Model_v8.xlsx")
print("="*70)
wb=openpyxl.load_workbook("Grifols_Model_v8.xlsx", data_only=True, read_only=True)
print("SHEETS:")
for s in wb.sheetnames:
    ws=wb[s]
    print(f"  - {s}  ({ws.max_row} x {ws.max_column})")

for target in ["IS","BS","Dashboard","Covenants"]:
    if target in wb.sheetnames:
        ws=wb[target]
        print("\n"+"#"*60)
        print(f"SHEET: {target}")
        print("#"*60)
        dump_headers(ws)
        dump_colA(ws)
wb.close()

print("\n"+"="*70)
print("ACTUALS: Grifols_Actuals_Reference.xlsx")
print("="*70)
wb2=openpyxl.load_workbook("Grifols_Actuals_Reference.xlsx", data_only=True, read_only=True)
print("SHEETS:")
for s in wb2.sheetnames:
    ws=wb2[s]
    print(f"  - {s}  ({ws.max_row} x {ws.max_column})")
for s in wb2.sheetnames:
    ws=wb2[s]
    print("\n"+"#"*60)
    print(f"ACTUALS SHEET: {s}")
    print("#"*60)
    for r in range(1, ws.max_row+1):
        cells=[]
        for c in range(1, ws.max_column+1):
            v=ws.cell(r,c).value
            if v is not None:
                s2=str(v)
                if len(s2)>18: s2=s2[:18]
                cells.append(f"{openpyxl.utils.get_column_letter(c)}={s2}")
        if cells:
            print(f"  r{r}: " + " | ".join(cells))
wb2.close()
print("\nDONE")
