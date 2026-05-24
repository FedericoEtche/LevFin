import openpyxl, zipfile, re

wb = openpyxl.load_workbook("Grifols_Model_v8.xlsx", data_only=True, read_only=True)
BS = wb["BS"]
def n(r,c):
    v=BS.cell(r,c).value
    return float(v) if isinstance(v,(int,float)) else v

print("=== LEASE-LIABILITY CLASSIFICATION CHECK ===")
for col,yr in [(6,"2024YE"),(10,"2025YE")]:
    bw_nc=n(39,col); bw_cur=n(47,col); ll_nc=n(40,col); ll_cur=n(48,col)
    print(f"\n{yr}:")
    print(f"  Model borrowings NC (r39)         {bw_nc:,.1f}")
    print(f"  Model lease liab NC (r40)         {ll_nc:,.1f}")
    print(f"  -> borrowings+lease NC            {bw_nc+ll_nc:,.1f}   (actuals NC financial liab: {'9490.6' if col==6 else '9091.0'})")
    print(f"  Model borrowings current (r47)    {bw_cur:,.1f}")
    print(f"  Model lease liab current (r48)    {ll_cur:,.1f}")
    print(f"  -> borrowings+lease current       {bw_cur+ll_cur:,.1f}   (actuals current financial liab: {'676.1' if col==6 else '552.0'})")
    print(f"  Total lease liability (NC+cur)    {ll_nc+ll_cur:,.1f}   (actuals memo lease line: {'1141.4' if col==6 else '1077.0'})")

# current/non-current 0.6 item at 2024YE
print("\n=== 2024YE current/non-current 0.6 boundary ===")
for r in [16,17,18,23,24,25]:
    print(f"  BS r{r} '{BS.cell(r,1).value}' = {n(r,6)}")
wb.close()

print("\n\n=== TIE-OUT MEMO (today) — text dump ===")
try:
    z=zipfile.ZipFile("Grifols_Model_TieOut_Memo_20260524.docx")
    xml=z.read("word/document.xml").decode("utf-8","ignore")
    # split into paragraphs
    paras=re.split(r"</w:p>", xml)
    out=[]
    for p in paras:
        texts=re.findall(r"<w:t[^>]*>(.*?)</w:t>", p)
        line="".join(texts)
        line=re.sub(r"&amp;","&",line); line=re.sub(r"&lt;","<",line); line=re.sub(r"&gt;",">",line)
        if line.strip(): out.append(line.strip())
    print(f"[{len(out)} non-empty paragraphs]\n")
    for ln in out[:160]:
        print(ln)
except Exception as e:
    print("ERR", e)
