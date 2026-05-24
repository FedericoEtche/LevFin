import openpyxl

wb = openpyxl.load_workbook("Grifols_Model_v8.xlsx", data_only=True, read_only=True)
IS = wb["IS"]; BS = wb["BS"]; DSH = wb["Dashboard"]; COV = wb["Covenants"]

def num(ws, r, c):
    v = ws.cell(r, c).value
    return float(v) if isinstance(v, (int, float)) else None

# IS columns: 2024 = C(3)..F(6); 2025 = G(7)..J(10)
def fy(ws, r, cols):
    vals = [num(ws, r, c) for c in cols]
    vals = [v for v in vals if v is not None]
    return sum(vals) if vals else None

C2024 = [3,4,5,6]; C2025 = [7,8,9,10]

# ---------- ACTUALS ----------
wa = openpyxl.load_workbook("Grifols_Actuals_Reference.xlsx", data_only=True, read_only=True)
AIS = wa["Quarterly IS"]; ABS = wa["Quarterly BS"]
# actuals IS: FY2024 col F(6), FY2025 col K(11)
def ais(r): return num(AIS, r, 6), num(AIS, r, 11)
# actuals BS: 2024YE col F(6), 2025YE col J(10)
def abs_(r): return num(ABS, r, 6), num(ABS, r, 10)

def line(label, m24, m25, a24, a25):
    def d(m,a):
        if m is None or a is None: return None
        return m-a
    def f(x): return "—" if x is None else f"{x:,.1f}"
    d24=d(m24,a24); d25=d(m25,a25)
    flag=""
    for dd in (d24,d25):
        if dd is not None and abs(dd)>0.5: flag="  <-- DIFF"
    print(f"{label:<34} | M24 {f(m24):>10} A24 {f(a24):>10} d {f(d24):>7} | M25 {f(m25):>10} A25 {f(a25):>10} d {f(d25):>7}{flag}")

print("="*150)
print("INCOME STATEMENT  (Model FY = sum of quarters)   vs   Actuals Reference")
print("="*150)
IS_MAP = [
 ("Revenue", 10, 5),
 ("Cost of sales", 12, 6),
 ("Gross profit", 13, 7),
 ("R&D", 16, 8),
 ("SG&A", 15, 9),
 ("EBITDA (model-derived)", 20, None),
 ("EBIT", 25, 13),
 ("Profit before tax", 31, 23),
 ("Income tax", 33, 24),
 ("Profit for the period", 39, 25),
 ("  o/w owners of parent", 42, 26),
 ("  o/w NCI", 41, 27),
]
for label, mr, ar in IS_MAP:
    m24 = fy(IS, mr, C2024); m25 = fy(IS, mr, C2025)
    a24, a25 = (ais(ar) if ar else (None, None))
    line(label, m24, m25, a24, a25)

# EBITDA reconciliation: EBIT + D&A
da24 = (fy(IS,22,C2024) or 0)+(fy(IS,23,C2024) or 0)
da25 = (fy(IS,22,C2025) or 0)+(fy(IS,23,C2025) or 0)
print(f"\n  [check] Model D&A (dep+amort) FY24 {da24:,.1f} / FY25 {da25:,.1f}")
print(f"  [check] Model EBIT+D&A FY24 {fy(IS,25,C2024)+da24:,.1f} / FY25 {fy(IS,25,C2025)+da25:,.1f}  (should equal model EBITDA)")

print("\n"+"="*150)
print("BALANCE SHEET  (Model 2024YE = col F / 2025YE = col J)   vs   Actuals Reference")
print("="*150)
# (label, model row, actuals row)
BS_MAP = [
 ("PP&E", 11, 10),
 ("Right-of-use assets", 12, 8),
 ("Goodwill", 13, 6),
 ("Other intangibles", 14, 7),
 ("Investments in associates", 15, 11),
 ("Deferred tax assets", 16, 12),
 ("Total non-current assets", 18, 14),
 ("Inventories", 21, 15),
 ("Trade & other receivables", 22, 16),
 ("Cash & equivalents", 24, 17),
 ("Total current assets", 25, 19),
 ("TOTAL ASSETS", 27, 20),
 ("Equity attrib. to owners", 34, 22),
 ("Non-controlling interests", 35, 23),
 ("Total equity", 36, 24),
 ("Borrowings — non-current", 39, 25),
 ("Borrowings — current", 47, 26),
 ("Deferred tax liabilities", 41, 28),
 ("Trade & other payables", 46, 29),
 ("Total liabilities", 54, 31),
 ("TOTAL EQUITY + LIAB", 55, 32),
]
for label, mr, ar in BS_MAP:
    m24 = num(BS, mr, 6); m25 = num(BS, mr, 10)
    a24, a25 = abs_(ar)
    line(label, m24, m25, a24, a25)

print("\n  [check] Model BS balance check (row57) 2024YE", num(BS,57,6), " 2025YE", num(BS,57,10))
print("  [check] Model hist restatement bridge (row53) 2024YE", num(BS,53,6), " 2025YE", num(BS,53,10))

print("\n"+"="*150)
print("LEVERAGE / COVENANT cross-check (Covenants sheet)")
print("="*150)
# total net leverage row26: 2025Q4 = J(10), 2026Q1 = K(11)
for r,lab in [(21,"Gross debt total"),(22,"Cash"),(23,"Net debt"),(25,"LTM EBITDA"),(26,"Total net leverage (x)"),(30,"Headroom (turns)")]:
    print(f"  {lab:<26} 2025Q4(J) {num(COV,r,10)}   2026Q1(K) {num(COV,r,11)}   2030Q4(AD) {num(COV,r,30)}")

print("\n"+"="*150)
print("DASHBOARD annual summary (FY) — rows 52-71, cols A-G")
print("="*150)
for r in range(54,72):
    cells=[]
    for c in range(1,8):
        v=DSH.cell(r,c).value
        if v is not None:
            s=str(v)
            if len(s)>16: s=s[:16]
            cells.append(f"{openpyxl.utils.get_column_letter(c)}={s}")
    if cells: print(f"  r{r}: "+" | ".join(cells))

wb.close(); wa.close()
print("\nDONE")
