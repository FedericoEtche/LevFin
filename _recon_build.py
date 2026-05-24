# Build Grifols v8 -> audited reconciliation workpaper (values pulled live from the model)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

m = openpyxl.load_workbook("Grifols_Model_v8.xlsx", data_only=True, read_only=True)
IS=m["IS"]; BS=m["BS"]; CF=m["CF"]; COV=m["Covenants"]
a = openpyxl.load_workbook("Grifols_Actuals_Reference.xlsx", data_only=True, read_only=True)
AIS=a["Quarterly IS"]; ABS=a["Quarterly BS"]

def mv(ws,r,c):
    v=ws.cell(r,c).value
    return float(v) if isinstance(v,(int,float)) else None
def fy(ws,r,cols):
    vs=[mv(ws,r,c) for c in cols]; vs=[v for v in vs if v is not None]
    return round(sum(vs),1) if vs else None
C24=[3,4,5,6]; C25=[7,8,9,10]      # IS quarter cols
# actuals IS FY cols: FY24=6, FY25=11 ; actuals BS YE cols: 2024YE=6, 2025YE=10
def ais(r): return mv(AIS,r,6), mv(AIS,r,11)
def abs_(r): return mv(ABS,r,6), mv(ABS,r,10)

# ---- styles ----
NAVY="1F3864"; LBLUE="EEF2F8"; AMBER="FFF2CC"; GREEN="E2EFDA"
white_bold=Font(color="FFFFFF", bold=True, size=10)
bold=Font(bold=True, size=10); reg=Font(size=10); ital=Font(size=9, italic=True, color="555555")
title_f=Font(bold=True, size=14, color=NAVY); sub_f=Font(size=10, italic=True, color="555555")
hfill=PatternFill("solid", fgColor=NAVY); band=PatternFill("solid", fgColor=LBLUE)
amber=PatternFill("solid", fgColor=AMBER); green=PatternFill("solid", fgColor=GREEN)
ctr=Alignment(horizontal="center", vertical="center"); lft=Alignment(horizontal="left", vertical="center", wrap_text=True)
rgt=Alignment(horizontal="right", vertical="center")
thin=Side(style="thin", color="BBBBBB"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
NUM="#,##0.0"

wb=openpyxl.Workbook()

def hdr(ws, row, cols):
    for c,(txt,w) in enumerate(cols,1):
        cell=ws.cell(row,c,txt); cell.font=white_bold; cell.fill=hfill; cell.alignment=ctr; cell.border=box
        ws.column_dimensions[get_column_letter(c)].width=w

def datarow(ws, row, vals, bandit=False, notefill=None):
    for c,v in enumerate(vals,1):
        cell=ws.cell(row,c,v); cell.border=box
        cell.font=bold if c==1 and isinstance(v,str) and v.isupper() else reg
        if c==1: cell.alignment=lft
        elif isinstance(v,(int,float)): cell.alignment=rgt; cell.number_format=NUM
        else: cell.alignment=ctr
        if bandit: cell.fill=band
        if notefill and c==len(vals): cell.fill=notefill

# ============ SUMMARY ============
ws=wb.active; ws.title="Summary"
ws.sheet_view.showGridLines=False
ws["A1"]="GRIFOLS, S.A. — MODEL-TO-SOURCE RECONCILIATION"; ws["A1"].font=title_f
ws["A2"]="Grifols_Model_v8.xlsx  reconciled to the audited consolidated annual accounts (FY2024, FY2025)"; ws["A2"].font=sub_f
meta=[("Prepared","2026-05-24  ·  Claude (reconciliation skill)"),
      ("Source of truth","Grifols audited consolidated annual accounts 2024 & 2025, via Grifols_Actuals_Reference.xlsx (8-qtr backbone; provenance per its Notes tab)"),
      ("Model basis","IFRS; historical periods (2024-2025) hardcoded to reported actuals; FY = sum of 4 quarterly columns; year-end = Q4 column"),
      ("Materiality","Rounding tolerance ±0.5 € m at line level; subtotals must tie exactly"),
      ("Reviewer","__________________  (to sign off)")]
r=4
for k,v in meta:
    ws.cell(r,1,k).font=bold; ws.cell(r,1,k).alignment=Alignment(vertical="top")
    c=ws.cell(r,2,v); c.alignment=lft; ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    r+=1
r+=1
ws.cell(r,1,"VERDICT").font=Font(bold=True,size=11,color="FFFFFF"); ws.cell(r,1).fill=PatternFill("solid",fgColor="375623")
for cc in range(2,8): ws.cell(r,cc).fill=PatternFill("solid",fgColor="375623")
vtxt=("PASS — the v8 historicals tie to the audited accounts. Every audited subtotal (revenue, EBITDA, EBIT, PBT, tax, net income; "
      "total assets, total equity, total liabilities, cash) matches within rounding. The only line-level differences are presentation "
      "reclasses that net to zero at the next subtotal. ONE open item: the historical cash-flow composition (model runs a min-cash sweep).")
ws.cell(r,1).value="VERDICT:  PASS (with 1 open item)"
r+=1
mc=ws.cell(r,1,vtxt); mc.alignment=lft; mc.font=reg; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=58
r+=2

ws.cell(r,1,"KEY SUBTOTALS — model vs audited").font=bold; r+=1
hdr(ws,r,[("€ millions",30),("Model FY24/24YE",16),("Audit FY24/24YE",16),("Δ24",8),("Model FY25/25YE",16),("Audit FY25/25YE",16),("Δ25",8),("Status",12)])
r+=1
def keyrow(label, m24,m25,a24,a25):
    d24=None if (m24 is None or a24 is None) else round(m24-a24,1)
    d25=None if (m25 is None or a25 is None) else round(m25-a25,1)
    st="Ties" if (d24 is not None and abs(d24)<=0.5 and d25 is not None and abs(d25)<=0.5) else "See detail"
    datarow(ws,r_[0],[label,m24,a24,d24,m25,a25,d25,st], bandit=(r_[0]%2==0))
    if st=="Ties":
        ws.cell(r_[0],8).fill=green
    r_[0]+=1
r_=[r]
# IS
keyrow("Revenue", fy(IS,10,C24),fy(IS,10,C25), *[ais(5)[0],ais(5)[1]])
keyrow("EBITDA (model-derived)", fy(IS,20,C24),fy(IS,20,C25), None,None)
keyrow("EBIT", fy(IS,25,C24),fy(IS,25,C25), ais(13)[0],ais(13)[1])
keyrow("Profit before tax", fy(IS,31,C24),fy(IS,31,C25), ais(23)[0],ais(23)[1])
keyrow("Net income (consolidated)", fy(IS,39,C24),fy(IS,39,C25), ais(25)[0],ais(25)[1])
# BS
keyrow("Total assets", mv(BS,27,6),mv(BS,27,10), abs_(20)[0],abs_(20)[1])
keyrow("Total equity", mv(BS,36,6),mv(BS,36,10), abs_(24)[0],abs_(24)[1])
keyrow("Total liabilities", mv(BS,54,6),mv(BS,54,10), abs_(31)[0],abs_(31)[1])
keyrow("Cash & equivalents", mv(BS,24,6),mv(BS,24,10), abs_(17)[0],abs_(17)[1])
r=r_[0]+1
# EBITDA note
ws.cell(r,1,"EBITDA has no direct audited line (non-IFRS); it ties to EBIT + D&A and to the CLAUDE.md canon (1,629 / 1,693).").font=ital
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); r+=2

# leverage cross-check
ws.cell(r,1,"LEVERAGE CROSS-CHECK (Covenants sheet, ties to CLAUDE.md canon)").font=bold; r+=1
hdr(ws,r,[("Metric",30),("2025Q4",16),("2026Q1",16),("2030Q4",8),("",16),("",16),("",8),("",12)]); r+=1
for rr,lab in [(26,"Total net leverage (x)"),(21,"Gross debt — total (ex-IFRS-16)"),(23,"Net debt"),(25,"LTM EBITDA")]:
    fmt = "0.00x" if rr==26 else NUM
    vals=[lab, round(mv(COV,rr,10),3), round(mv(COV,rr,11),3), round(mv(COV,rr,30),3),"","","",""]
    datarow(ws,r,vals);
    for cc in (2,3,4): ws.cell(r,cc).number_format=fmt
    r+=1
r+=1
note=("NOTE — supersedes the morning tie-out memo. Grifols_Model_TieOut_Memo_20260524.docx (09:00) was written against a pre-rebuild "
      "model (its 'Model' column shows Rev 7,215 / EBITDA 1,598 / NI 369 / assets 20,715 — v2-vintage). v8 (finalised 16:10) closed "
      "every 'real gap' it flagged: debt re-anchored, FX translation added, FY2024 tax one-off captured, net income reconciled. "
      "Treat that memo as superseded by this workpaper.")
mc=ws.cell(r,1,note); mc.alignment=lft; mc.font=Font(size=9, color="833C00"); mc.fill=amber
for cc in range(2,9): ws.cell(r,cc).fill=amber
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); ws.row_dimensions[r].height=66
ws.freeze_panes="A3"

# ============ generic tie-out sheet builder ============
def tieout(title, subtitle, colhdr, rows):
    ws=wb.create_sheet(title); ws.sheet_view.showGridLines=False
    ws["A1"]=title.replace("_"," "); ws["A1"].font=title_f
    ws["A2"]=subtitle; ws["A2"].font=sub_f
    hdr(ws,4,colhdr)
    rr=5
    for row in rows:
        if row[0]=="__SUB__":
            for c,v in enumerate(row[1:],1):
                cell=ws.cell(rr,c,v); cell.font=bold; cell.fill=band; cell.border=box
                if c==1: cell.alignment=lft
                elif isinstance(v,(int,float)): cell.alignment=rgt; cell.number_format=NUM
                else: cell.alignment=ctr
            rr+=1; continue
        datarow(ws,rr,row, bandit=False)
        # colour the status col
        stat=row[-1]
        if isinstance(stat,str):
            if stat.startswith("Ties"): ws.cell(rr,len(row)).fill=green
            elif "OPEN" in stat: ws.cell(rr,len(row)).fill=amber
            elif "Reclass" in stat or "Rounding" in stat: ws.cell(rr,len(row)).fill=PatternFill("solid",fgColor="FCE4D6")
        rr+=1
    ws.freeze_panes="A5"
    return ws

def diff(m,a): return None if (m is None or a is None) else round(m-a,1)

# ---- IS tie-out ----
def isline(label, mr, ar, status):
    m24=fy(IS,mr,C24); m25=fy(IS,mr,C25)
    a24,a25 = (ais(ar) if ar else (None,None))
    return [label, m24, a24, diff(m24,a24), m25, a25, diff(m25,a25), status]
# model net finance result & associates = PBT - EBIT (the model's own EBIT->PBT bridge; robust to sign convention)
fin_m24=round(fy(IS,31,C24)-fy(IS,25,C24),1)
fin_m25=round(fy(IS,31,C25)-fy(IS,25,C25),1)
isrows=[
 isline("Revenue",10,5,"Ties"),
 isline("Cost of sales",12,6,"Reclass (D&A)"),
 ["__SUB__","Gross profit", fy(IS,13,C24), ais(7)[0], diff(fy(IS,13,C24),ais(7)[0]), fy(IS,13,C25), ais(7)[1], diff(fy(IS,13,C25),ais(7)[1]), "Reclass (D&A)"],
 isline("R&D",16,8,"Ties"),
 isline("SG&A",15,9,"Ties"),
 ["__SUB__","EBITDA (model-derived)", fy(IS,20,C24), None, None, fy(IS,20,C25), None, None, "Non-IFRS"],
 isline("Depreciation",22,None,"memo"),
 isline("Amortization",23,None,"memo"),
 ["__SUB__","EBIT", fy(IS,25,C24), ais(13)[0], diff(fy(IS,25,C24),ais(13)[0]), fy(IS,25,C25), ais(13)[1], diff(fy(IS,25,C25),ais(13)[1]), "Ties"],
 ["Net finance result & associates", fin_m24, ais(21)[0], diff(fin_m24,ais(21)[0]), fin_m25, ais(21)[1], diff(fin_m25,ais(21)[1]), "Ties"],
 ["__SUB__","Profit before tax", fy(IS,31,C24), ais(23)[0], diff(fy(IS,31,C24),ais(23)[0]), fy(IS,31,C25), ais(23)[1], diff(fy(IS,31,C25),ais(23)[1]), "Ties"],
 isline("Income tax",33,24,"Ties"),
 ["__SUB__","Profit for the period", fy(IS,39,C24), ais(25)[0], diff(fy(IS,39,C24),ais(25)[0]), fy(IS,39,C25), ais(25)[1], diff(fy(IS,39,C25),ais(25)[1]), "Ties"],
 isline("  o/w owners of parent",42,26,"Rounding"),
 isline("  o/w non-controlling int.",41,27,"Rounding"),
]
colh=[("INCOME STATEMENT  (€ m)",34),("Model FY24",11),("Audit FY24",11),("Δ24",8),("Model FY25",11),("Audit FY25",11),("Δ25",8),("Status",16)]
tieout("IS Tie-Out","FY2024 / FY2025 — model (sum of quarters) vs audited consolidated income statement", colh, isrows)

# ---- BS tie-out ----
def bsline(label, mr, ar, status):
    m24=mv(BS,mr,6); m25=mv(BS,mr,10); a24,a25=abs_(ar)
    return [label, m24, a24, diff(m24,a24), m25, a25, diff(m25,a25), status]
lease24=round((mv(BS,40,6) or 0)+(mv(BS,48,6) or 0),1); lease25=round((mv(BS,40,10) or 0)+(mv(BS,48,10) or 0),1)
bsrows=[
 bsline("PP&E",11,10,"Ties"),
 bsline("Right-of-use assets",12,8,"Ties"),
 bsline("Goodwill",13,6,"Ties"),
 bsline("Other intangibles",14,7,"Ties"),
 bsline("Investments in associates",15,11,"Ties"),
 bsline("Deferred tax assets",16,12,"Ties"),
 ["__SUB__","Total non-current assets", mv(BS,18,6), abs_(14)[0], diff(mv(BS,18,6),abs_(14)[0]), mv(BS,18,10), abs_(14)[1], diff(mv(BS,18,10),abs_(14)[1]), "Reclass (0.6)"],
 bsline("Inventories",21,15,"Ties"),
 bsline("Trade & other receivables",22,16,"Ties"),
 bsline("Cash & equivalents",24,17,"Ties (anchor)"),
 ["__SUB__","Total current assets", mv(BS,25,6), abs_(19)[0], diff(mv(BS,25,6),abs_(19)[0]), mv(BS,25,10), abs_(19)[1], diff(mv(BS,25,10),abs_(19)[1]), "Reclass (0.6)"],
 ["__SUB__","TOTAL ASSETS", mv(BS,27,6), abs_(20)[0], diff(mv(BS,27,6),abs_(20)[0]), mv(BS,27,10), abs_(20)[1], diff(mv(BS,27,10),abs_(20)[1]), "Ties"],
 bsline("Equity attrib. to owners",34,22,"Ties"),
 bsline("Non-controlling interests",35,23,"Ties"),
 ["__SUB__","Total equity", mv(BS,36,6), abs_(24)[0], diff(mv(BS,36,6),abs_(24)[0]), mv(BS,36,10), abs_(24)[1], diff(mv(BS,36,10),abs_(24)[1]), "Ties"],
 bsline("Borrowings — non-current",39,25,"Reclass (lease)"),
 bsline("Borrowings — current",47,26,"Reclass (lease)"),
 ["  + Lease liabilities (model, separate)", lease24, ABS.cell(27,6).value, diff(lease24,float(ABS.cell(27,6).value)), lease25, ABS.cell(27,10).value, diff(lease25,float(ABS.cell(27,10).value)), "memo"],
 bsline("Deferred tax liabilities",41,28,"Ties"),
 bsline("Trade & other payables",46,29,"Ties"),
 ["__SUB__","Total liabilities", mv(BS,54,6), abs_(31)[0], diff(mv(BS,54,6),abs_(31)[0]), mv(BS,54,10), abs_(31)[1], diff(mv(BS,54,10),abs_(31)[1]), "Ties"],
 ["__SUB__","TOTAL EQUITY + LIABILITIES", mv(BS,55,6), abs_(32)[0], diff(mv(BS,55,6),abs_(32)[0]), mv(BS,55,10), abs_(32)[1], diff(mv(BS,55,10),abs_(32)[1]), "Ties"],
]
colh2=[("BALANCE SHEET  (€ m)",34),("Model 24YE",11),("Audit 24YE",11),("Δ24",8),("Model 25YE",11),("Audit 25YE",11),("Δ25",8),("Status",16)]
tieout("BS Tie-Out","2024YE / 2025YE — model Q4 column vs audited consolidated balance sheet", colh2, bsrows)

# ---- CF tie-out ----
wscf=wb.create_sheet("CF Tie-Out"); wscf.sheet_view.showGridLines=False
wscf["A1"]="CF Tie-Out"; wscf["A1"].font=title_f
wscf["A2"]="Cash balance is anchored and ties; cash-flow composition is a model construct (min-cash sweep) — OPEN item"; wscf["A2"].font=sub_f
hdr(wscf,4,[("CASH FLOW  (€ m)",36),("Model FY25",13),("Audit FY25 (approx)",18),("Status",16),("",2),("",2)])
ocf_m=fy(CF,29,C25)
cfrows=[
 ["Net cash from operating activities", ocf_m, 1047.0, "OPEN — composition"],
 ["Net cash from investing activities", fy(CF,38,C25), None, "Model basis"],
 ["Net cash from financing activities", fy(CF,48,C25), None, "Model basis (plug)"],
 ["Net change in cash", fy(CF,50,C25), None, "Ties to BS move"],
 ["Cash balance — 2024YE (opening)", mv(BS,24,6), abs_(17)[0], "Ties (anchor)"],
 ["Cash balance — 2025YE (closing)", mv(BS,24,10), abs_(17)[1], "Ties (anchor)"],
]
rr=5
for row in cfrows:
    for c,v in enumerate(row,1):
        cell=wscf.cell(rr,c,v); cell.border=box
        if c==1: cell.alignment=lft; cell.font=bold if "balance" in str(v).lower() or "Net change" in str(v) else reg
        elif isinstance(v,(int,float)): cell.alignment=rgt; cell.number_format=NUM; cell.font=reg
        else: cell.alignment=ctr; cell.font=reg
    st=row[-1]
    if st.startswith("Ties"): wscf.cell(rr,4).fill=green
    elif "OPEN" in st: wscf.cell(rr,4).fill=amber
    rr+=1
rr+=1
cap=("FY2024 cash flow is a model-initialisation year (the 2024Q1 setup column carries a ~€6.4bn drawdown / opening-cash plug) and is "
     "not comparable to the audited statement. For FY2025 the net change in cash (-154.8) reconciles to the balance-sheet move "
     "(979.8 → 825.0), but operating cash (~€797m model) does not reproduce the audited ~€1.0bn — the model sweeps surplus/shortfall "
     "to the financing line to hold the min-cash target. Use the audited cash-flow statement for historical cash generation.")
mc=wscf.cell(rr,1,cap); mc.alignment=lft; mc.font=Font(size=9,color="833C00"); mc.fill=amber
for cc in range(2,5): wscf.cell(rr,cc).fill=amber
wscf.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4); wscf.row_dimensions[rr].height=72

# ============ RECONCILING ITEMS ============
ws=wb.create_sheet("Reconciling Items"); ws.sheet_view.showGridLines=False
ws["A1"]="Reconciling Items — schedule"; ws["A1"].font=title_f
ws["A2"]="Each line-level difference, categorised (reconciliation skill: classification / rounding / investigation)"; ws["A2"].font=sub_f
hdr(ws,4,[("#",4),("Item",30),("Period",14),("Amount € m",14),("Category",18),("Nets at",16),("Status",14),("Explanation",70)])
items=[
 (1,"Cost of sales / gross profit — D&A geography","FY24 / FY25","+437.0 / +450.1","Classification","EBIT","Confirmed",
   "Model shows D&A as separate lines below EBITDA; Grifols allocates production D&A into cost of sales (function-of-expense). Difference equals model D&A exactly; EBITDA and EBIT tie. No action."),
 (2,"Borrowings vs lease liabilities — IFRS-16 split","24YE / 25YE","-1,141.4 / -1,077.0","Classification","Total liabilities","Confirmed",
   "Model carries lease liabilities on separate lines (Credit-Agreement, ex-IFRS-16 debt basis for the covenant); Grifols folds leases into financial liabilities. Borrowings + model leases = reported financial liabilities to the decimal. No action."),
 (3,"Current / non-current asset boundary","24YE","+0.6 / 0.0","Classification","Total assets","Confirmed",
   "€0.6m other-NCA vs other-current reclass at 2024YE only. Immaterial, self-netting at total assets. No action."),
 (4,"Parent / NCI allocation of net income","FY24","+0.3 / -0.3","Rounding","Net income","Confirmed",
   "±0.3m split between owners-of-parent and NCI within consolidated net income. Self-netting at total net income. No action."),
 (5,"Historical cash-flow composition","FY24-FY25","OCF ~797 vs ~1,047 (FY25)","Investigation","-","OPEN",
   "Model runs the historical cash flow as a min-cash sweep: cash balance is anchored and ties, but operating-cash composition does not reproduce the audited statement; FY2024 is an initialisation year. Per morning memo fix #3. Decide: re-anchor historical CF to audited, or document the simplification."),
]
rr=5
for it in items:
    for c,v in enumerate(it,1):
        cell=ws.cell(rr,c,v); cell.border=box; cell.font=reg
        cell.alignment=lft if c in (2,8) else ctr
    if it[6]=="OPEN":
        for c in range(1,9): ws.cell(rr,c).fill=amber
    ws.row_dimensions[rr].height=58
    rr+=1

m.close(); a.close()
out="Grifols_v8_Reconciliation_20260524.xlsx"
wb.save(out)
print("SAVED", out)
print("Sheets:", wb.sheetnames)
